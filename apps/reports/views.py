import csv
import json
import os
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from io import TextIOWrapper
from typing import Optional

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db.models import Avg, Count, DurationField, ExpressionWrapper, F, Q
from django.http import HttpRequest, HttpResponse, HttpResponseNotAllowed, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.dateparse import parse_datetime
from django.utils.text import slugify
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET, require_http_methods

from apps.accounts.models import AuditLog, UserProfile
from apps.accounts.utils import notify_report_classified, notify_report_reassigned
from apps.ai_classifier.classifier import classify_report
from apps.notifications.senders import send_status_change_email
from apps.notifications.services import send_report_created_email, send_report_status_changed_email

from .duplicate_detection import find_potential_duplicate
from .forms import (
    AdminUserUpdateForm,
    AdminUserCreateForm,
    ReportCategoryForm,
    ReportCreateForm,
    ReportStatusUpdateForm,
    ReportSubmissionForm,
    SectorForm,
)
from .models import MUNICIPALITY_CHOICES, Report, ReportCategory, ReportStatusHistory, Sector


MAX_REPORTS_PER_24H = 10
REPORT_WINDOW_HOURS = 24


SEARCH_PARAMS = (
    "category", "status", "sector", "priority",
    "from", "to", "keyword",
    "lat_min", "lat_max", "lng_min", "lng_max",
    "page",
)

REPORT_EXPORT_COLUMNS = [
    "id",
    "description",
    "category",
    "priority",
    "status",
    "sector",
    "location",
    "latitude",
    "longitude",
    "municipality",
    "citizen",
    "created_at",
    "updated_at",
    "status_changed_at",
]


def _parse_iso_date(value):
    """Return a ``date`` parsed from ISO-8601 input, or ``None`` on failure."""
    try:
        return date.fromisoformat(value)
    except (TypeError, ValueError):
        return None


def _parse_decimal(value):
    """Return a ``Decimal`` or ``None`` if the input is absent/invalid."""
    if value is None or value == "":
        return None
    try:
        return Decimal(value)
    except (InvalidOperation, TypeError, ValueError):
        return None


def _build_report_filters(request):
    """Translate GET parameters into a ``Q()`` expression."""
    filters = Q()

    for param, field in (
        ("category", "category"),
        ("status", "status"),
        ("sector", "sector"),
        ("priority", "priority"),
    ):
        value = request.GET.get(param)
        if value:
            filters &= Q(**{field: value})

    from_date = _parse_iso_date(request.GET.get("from"))
    if from_date:
        filters &= Q(created_at__date__gte=from_date)

    to_date = _parse_iso_date(request.GET.get("to"))
    if to_date:
        filters &= Q(created_at__date__lte=to_date)

   
    keyword = request.GET.get("keyword")
    if keyword:
        clean_keyword = keyword.strip()
        if clean_keyword.upper().startswith("ПРЈ-"):
            clean_keyword = clean_keyword[4:]
        matching_slugs = [
            slug for slug, label in MUNICIPALITY_CHOICES
            if keyword.lower() in label.lower()
        ]
        filters &= (
            Q(description__icontains=keyword) |
            Q(id__icontains=clean_keyword) |
            Q(municipality__in=matching_slugs) |
            Q(category__icontains=keyword)
        )
    for param, lookup in (
        ("lat_min", "latitude__gte"),
        ("lat_max", "latitude__lte"),
        ("lng_min", "longitude__gte"),
        ("lng_max", "longitude__lte"),
    ):
        value = _parse_decimal(request.GET.get(param))
        if value is not None:
            filters &= Q(**{lookup: value})

    return filters


def _is_json_request(request):
    """True if caller prefers JSON (via Accept header or ``?format=json``)."""
    accept = request.headers.get("Accept", "")
    return "application/json" in accept.lower() or request.GET.get("format") == "json"


def _serialize_reports_page(page):
    return {
        "count": page.paginator.count,
        "num_pages": page.paginator.num_pages,
        "page": page.number,
        "results": [_serialize_report(report) for report in page.object_list],
    }

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

    
def home(request):
    """Landing page (no params) or paginated search endpoint (with params)."""
    should_filter = _is_json_request(request) or any(
        request.GET.get(param) for param in SEARCH_PARAMS
    )

    if not should_filter:
        if request.user.is_authenticated and request.user.is_superuser:
            return redirect("/dashboard/")
        return render(request, "reports/home.html")

    if not request.user.is_authenticated:
        if _is_json_request(request):
            return JsonResponse({"detail": "Authentication required."}, status=401)
        return redirect(f"{reverse('login')}?next={request.get_full_path()}")

    filters = _build_report_filters(request)
    queryset = Report.objects.filter(filters).order_by("-created_at")

    paginator = Paginator(queryset, 20)
    page_number = request.GET.get("page", 1)
    try:
        page_obj = paginator.page(page_number)
    except (PageNotAnInteger, EmptyPage):
        page_obj = paginator.page(1)

    if _is_json_request(request):
        return JsonResponse(_serialize_reports_page(page_obj), status=200)

    return render(
        request,
        "reports/search_results.html",
        {"reports_page": page_obj, "query": request.GET},
    )

def _as_float(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _is_officer(user) -> bool:
    return user.groups.filter(name__in=["officer", "officers"]).exists()


def _serialize_report(report: Report) -> dict:
    return {
        "id": report.id,
        "description": report.description,
        "category": report.category,
        "category_display": report.get_category_display(),
        "priority": report.priority,
        "priority_display": report.get_priority_display(),
        "status": report.status,
        "status_display": report.get_status_display(),
        "sector": report.sector,
        "sector_display": report.get_sector_display(),
        "latitude": _as_float(report.latitude),
        "longitude": _as_float(report.longitude),
        "created_at": report.created_at.isoformat(),
        "updated_at": report.updated_at.isoformat(),
        "status_changed_at": report.status_changed_at.isoformat() if report.status_changed_at else None,
        "detail_url": reverse("report_detail", args=[report.id]),
    }


def _apply_ai_classification(report: Report) -> None:
    result = classify_report(report.description)
    report.category = result.get("category", "other")
    report.priority = result.get("priority", "normal")
    report.sector = result.get("sector", "admin")
    report.status = result.get("status", "new")
    report.ai_processed = True
    report.status_changed_at = timezone.now()


def _parse_json_body(request):
    if "application/json" not in request.content_type:
        return None
    try:
        return json.loads((request.body or b"{}").decode("utf-8"))
    except (json.JSONDecodeError, UnicodeDecodeError):
        return {}


def _filter_queryset(request, queryset):
    category = (request.GET.get("category") or "").strip().lower()
    status = (request.GET.get("status") or "").strip().lower()
    sector = (request.GET.get("sector") or "").strip().lower()
    priority = (request.GET.get("priority") or "").strip().lower()
    municipality = (request.GET.get("municipality") or "").strip().lower()
    keyword = (request.GET.get("keyword") or "").strip()
    location = (request.GET.get("location") or "").strip()
    date_from = (request.GET.get("from") or "").strip()
    date_to = (request.GET.get("to") or "").strip()

    if category:
        queryset = queryset.filter(category=category)
    if status:
        queryset = queryset.filter(status=status)
    if sector:
        queryset = queryset.filter(sector=sector)
    if priority:
        queryset = queryset.filter(priority=priority)
    if municipality:
        queryset = queryset.filter(municipality=municipality)
    if keyword:
        queryset = queryset.filter(description__icontains=keyword)
    if location:
        queryset = queryset.filter(description__icontains=location)

    if date_from:
        try:
            dt_from = datetime.fromisoformat(date_from)
            queryset = queryset.filter(created_at__gte=dt_from)
        except ValueError:
            pass

    if date_to:
        try:
            dt_to = datetime.fromisoformat(date_to)
            queryset = queryset.filter(created_at__lte=dt_to)
        except ValueError:
            pass

    return queryset


def _filtered_admin_reports(request):
    return _filter_queryset(request, Report.objects.select_related("citizen").order_by("-created_at"))


def _format_report_row(report):
    latitude = str(report.latitude)
    longitude = str(report.longitude)
    return [
        report.id,
        report.description,
        report.category,
        report.priority,
        report.status,
        report.sector,
        f"{latitude},{longitude}",
        latitude,
        longitude,
        report.municipality,
        report.citizen.username,
        report.created_at.isoformat(),
        report.updated_at.isoformat(),
        report.status_changed_at.isoformat() if report.status_changed_at else "",
    ]


def _clean_import_header(value):
    return str(value or "").strip().lower().replace(" ", "_")


def _parse_import_datetime(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return timezone.make_aware(value) if timezone.is_naive(value) else value
    if isinstance(value, date):
        return timezone.make_aware(datetime.combine(value, datetime.min.time()))

    parsed = parse_datetime(str(value).strip())
    if parsed:
        return timezone.make_aware(parsed) if timezone.is_naive(parsed) else parsed
    parsed_date = _parse_iso_date(str(value).strip())
    if parsed_date:
        return timezone.make_aware(datetime.combine(parsed_date, datetime.min.time()))
    return None


def _parse_import_location(row):
    latitude = row.get("latitude") or row.get("lat")
    longitude = row.get("longitude") or row.get("lng") or row.get("lon")
    if (not latitude or not longitude) and row.get("location"):
        pieces = [piece.strip() for piece in str(row["location"]).split(",")]
        if len(pieces) >= 2:
            latitude, longitude = pieces[0], pieces[1]
    return _parse_decimal(latitude), _parse_decimal(longitude)


def _read_import_rows(uploaded_file):
    extension = os.path.splitext(uploaded_file.name)[1].lower()
    if extension == ".csv":
        text_file = TextIOWrapper(uploaded_file.file, encoding="utf-8-sig", newline="")
        return list(csv.DictReader(text_file))

    if extension in {".xlsx", ".xlsm"} and OPENPYXL_AVAILABLE:
        workbook = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [_clean_import_header(value) for value in rows[0]]
        return [
            {headers[index]: value for index, value in enumerate(row) if index < len(headers)}
            for row in rows[1:]
            if any(value not in (None, "") for value in row)
        ]

    raise ValueError("Поддржани формати се CSV и XLSX.")


def _import_reports_from_rows(rows, user):
    valid_categories = {choice[0] for choice in Report.CATEGORY_CHOICES} | set(
        ReportCategory.objects.values_list("key", flat=True)
    )
    valid_priorities = {choice[0] for choice in Report.PRIORITY_CHOICES}
    valid_statuses = {choice[0] for choice in Report.STATUS_CHOICES}
    active_sector_keys = set(Sector.objects.values_list("key", flat=True)) | {choice[0] for choice in Report.SECTOR_CHOICES}
    valid_municipalities = {choice[0] for choice in MUNICIPALITY_CHOICES}

    inserted = 0
    skipped_duplicates = []
    invalid_rows = []

    for index, raw_row in enumerate(rows, start=2):
        row = {_clean_import_header(key): value for key, value in raw_row.items()}
        errors = []
        report_id = str(row.get("id") or "").strip()
        description = str(row.get("description") or "").strip()
        category = str(row.get("category") or "").strip().lower()
        priority = str(row.get("priority") or "").strip().lower()
        status = str(row.get("status") or "").strip().lower()
        sector = str(row.get("sector") or "").strip().lower()
        municipality = str(row.get("municipality") or "").strip().lower()
        latitude, longitude = _parse_import_location(row)

        report_pk = None
        if report_id:
            try:
                report_pk = int(report_id)
            except ValueError:
                errors.append("invalid id")

        if report_pk and Report.objects.filter(pk=report_pk).exists():
            skipped_duplicates.append(report_id)
            continue
        if not description:
            errors.append("missing description")
        if category not in valid_categories:
            errors.append("invalid category")
        if priority not in valid_priorities:
            errors.append("invalid priority")
        if status not in valid_statuses:
            errors.append("invalid status")
        if sector not in active_sector_keys:
            errors.append("invalid sector")
        if municipality and municipality not in valid_municipalities:
            errors.append("invalid municipality")
        if latitude is None or longitude is None:
            errors.append("invalid location")

        created_at = _parse_import_datetime(row.get("created_at"))
        updated_at = _parse_import_datetime(row.get("updated_at"))
        status_changed_at = _parse_import_datetime(row.get("status_changed_at"))

        if errors:
            invalid_rows.append({"row": index, "reason": ", ".join(errors)})
            continue

        report = Report(
            citizen=user,
            description=description,
            latitude=latitude,
            longitude=longitude,
            category=category,
            priority=priority,
            status=status,
            sector=sector,
            municipality=municipality,
            status_changed_at=status_changed_at,
            ai_processed=True,
        )
        if report_pk:
            report.id = report_pk
        report.save()

        update_fields = []
        if created_at:
            report.created_at = created_at
            update_fields.append("created_at")
        if updated_at:
            report.updated_at = updated_at
            update_fields.append("updated_at")
        if update_fields:
            Report.objects.filter(pk=report.pk).update(**{field: getattr(report, field) for field in update_fields})
        inserted += 1

    return inserted, skipped_duplicates, invalid_rows


def _visible_reports_for_user(request):
    queryset = Report.objects.select_related("citizen").order_by("-created_at")
    if _is_admin_user(request.user):
        return queryset

    if _is_officer(request.user):
        profile = UserProfile.objects.filter(user=request.user).first()
        if profile and profile.sector:
            queryset = queryset.filter(sector=profile.sector)
            if profile.municipality:
                queryset = queryset.filter(municipality=profile.municipality)
            return queryset
        return queryset.none()

    return queryset.filter(citizen=request.user)


def _can_view_report(user, report: Report) -> bool:
    if user.is_superuser:
        return True
    if report.citizen_id == user.id:
        return True
    if _is_officer(user):
        profile = UserProfile.objects.filter(user=user).first()
        if not profile or not profile.sector or profile.sector != report.sector:
            return False
        return not profile.municipality or profile.municipality == report.municipality
    return False


def _is_submission_rate_limited(user) -> bool:
    cutoff = timezone.now() - timedelta(hours=REPORT_WINDOW_HOURS)
    recent_reports_count = Report.objects.filter(citizen=user, created_at__gte=cutoff).count()
    return recent_reports_count >= MAX_REPORTS_PER_24H


def _remaining_reports_quota(user) -> int:
    cutoff = timezone.now() - timedelta(hours=REPORT_WINDOW_HOURS)
    recent_reports_count = Report.objects.filter(citizen=user, created_at__gte=cutoff).count()
    return max(0, MAX_REPORTS_PER_24H - recent_reports_count)


def _log_status_transition(report: Report, from_status: Optional[str], to_status: str, changed_by=None, note: str = "") -> None:
    ReportStatusHistory.objects.create(
        report=report,
        from_status=from_status or "",
        to_status=to_status,
        changed_by=changed_by,
        note=note,
    )

def _is_admin_user(user: User) -> bool:
    """Allow dashboard access to superusers and admin group users."""
    if not user.is_authenticated:
        return False
    return user.is_superuser or user.groups.filter(name__in=["admin", "administrators"]).exists()


def _admin_only() -> callable:
    """Decorator for admin-only endpoints."""
    return user_passes_test(_is_admin_user)


def _write_audit_log(request, action, target_model, target_id, details):
    """Persist admin action to system log."""
    AuditLog.objects.create(
        user=request.user,
        action=action,
        target_model=target_model,
        target_id=target_id,
        details=details,
    )


def _build_unique_key(model, raw_name):
    """Generate a unique slug key for settings entities based on name."""
    base_key = slugify(raw_name)[:45] or "item"
    key = base_key
    counter = 2

    while model.objects.filter(key=key).exists():
        suffix = f"-{counter}"
        key = f"{base_key[:50 - len(suffix)]}{suffix}"
        counter += 1

    return key


@login_required
def dashboard(request):
    """Post-login landing: admin panel for admins, role-appropriate redirect otherwise."""
    if not _is_admin_user(request.user):
        if _is_officer(request.user):
            return redirect("officer_panel")
        return redirect("home")

    total_reports = Report.objects.count()
    active_users = User.objects.filter(is_active=True).count()
    resolved_reports = Report.objects.filter(status="resolved").count()
    resolve_rate = round((resolved_reports / total_reports) * 100, 1) if total_reports else 0

    avg_resolution_seconds = (
        Report.objects.filter(status="resolved", status_changed_at__isnull=False)
        .annotate(
            resolution_duration=ExpressionWrapper(
                F("status_changed_at") - F("created_at"),
                output_field=DurationField(),
            )
        )
        .aggregate(avg_duration=Avg("resolution_duration"))
    )
    avg_days = 0
    avg_duration = avg_resolution_seconds.get("avg_duration")
    if avg_duration:
        avg_days = round(avg_duration.total_seconds() / 86400, 1)

    category_counts = list(
        Report.objects.values("category").annotate(total=Count("id")).order_by("category")
    )
    status_counts = list(
        Report.objects.values("status").annotate(total=Count("id")).order_by("status")
    )
    priority_counts = list(
        Report.objects.values("priority").annotate(total=Count("id")).order_by("priority")
    )
    sector_counts = list(
        Report.objects.values("sector").annotate(total=Count("id")).order_by("sector")
    )

    missing_profile_users = User.objects.filter(profile__isnull=True)
    for listed_user in missing_profile_users:
        UserProfile.objects.get_or_create(user=listed_user)

    users_queryset = User.objects.select_related("profile").order_by("username")
    role_filter = request.GET.get("role", "")
    valid_roles = {value for value, _ in UserProfile.ROLE_CHOICES}
    if role_filter not in valid_roles:
        role_filter = ""
    users = users_queryset.filter(profile__role=role_filter) if role_filter else users_queryset
    role_filter_cards = [
        {"value": "", "label": "Сите", "count": users_queryset.count()},
        {
            "value": "citizen",
            "label": "Граѓани",
            "count": users_queryset.filter(profile__role="citizen").count(),
        },
        {
            "value": "officer",
            "label": "Работници",
            "count": users_queryset.filter(profile__role="officer").count(),
        },
        {
            "value": "admin",
            "label": "Админи",
            "count": users_queryset.filter(profile__role="admin").count(),
        },
    ]
    categories = ReportCategory.objects.order_by("name")
    sectors = Sector.objects.order_by("name")
    logs = AuditLog.objects.select_related("user").order_by("-timestamp")[:20]
    
    # Unclassified reports - those with "other" category or "unclassified" status
    unclassified_reports = Report.objects.filter(
        Q(category="other") | Q(status="unclassified")
    ).select_related("citizen").order_by("-created_at")[:50]

    pending_duplicate_reports = (
        Report.objects.filter(duplicate_verdict="pending")
        .exclude(duplicate_of__isnull=True)
        .select_related("citizen", "duplicate_of")
        .order_by("-created_at")[:100]
    )
    pending_duplicate_count = Report.objects.filter(duplicate_verdict="pending").exclude(duplicate_of__isnull=True).count()
    
    # For classification form: exclude "Друго" (other) from categories - it's the unclassified marker
    active_categories_for_classification = list(
        ReportCategory.objects.filter(is_active=True).exclude(key="other").values_list('key', 'name')
    )
    bulk_municipality = request.GET.get("bulk_municipality", "")
    bulk_sector = request.GET.get("bulk_sector", "")
    bulk_queryset = Report.objects.filter(status="resolved").exclude(citizen__email="")
    if bulk_municipality:
        bulk_queryset = bulk_queryset.filter(municipality=bulk_municipality)
    if bulk_sector:
        bulk_queryset = bulk_queryset.filter(sector=bulk_sector)
    default_bulk_subject = "Известување од Безбеден Град"
    default_bulk_message = (
        "Почитувани,\n\n"
        "Ве информираме дека пријавите што одговараат на избраните филтри се обработени. "
        "Ви благодариме што придонесувате за побезбедна заедница.\n\n"
        "Со почит,\nТимот на Безбеден Град"
    )

    context = {
        "active_tab": request.GET.get("tab", "users"),
        "stats": {
            "total_reports": total_reports,
            "active_users": active_users,
            "resolve_rate": resolve_rate,
            "avg_days": avg_days,
            "open_reports": Report.objects.exclude(status__in=["resolved", "rejected", "withdrawn"]).count(),
            "high_priority_reports": Report.objects.filter(priority="urgent").count(),
            "unclassified_reports": Report.objects.filter(Q(category="other") | Q(status="unclassified")).count(),
        },
        "category_counts": category_counts,
        "status_counts": status_counts,
        "priority_counts": priority_counts,
        "sector_counts": sector_counts,
        "users": users,
        "selected_user_role": role_filter,
        "role_filter_cards": role_filter_cards,
        "categories": categories,
        "sectors": sectors,
        "logs": logs,
        "unclassified_reports": unclassified_reports,
        "unclassified_count": Report.objects.filter(Q(category="other") | Q(status="unclassified")).count(),
        "pending_duplicate_reports": pending_duplicate_reports,
        "pending_duplicate_count": pending_duplicate_count,
        "category_form": ReportCategoryForm(),
        "sector_form": SectorForm(),
        "user_form": AdminUserCreateForm(),
        "role_choices": AdminUserCreateForm.ROLE_CHOICES,
        "sector_choices": list(Sector.objects.filter(is_active=True).values_list('key', 'name')),
        "category_choices": active_categories_for_classification,
        "status_choices": Report.STATUS_CHOICES,
        "priority_choices": Report.PRIORITY_CHOICES,
        "municipality_choices": MUNICIPALITY_CHOICES,
        "export_columns": REPORT_EXPORT_COLUMNS,
        "bulk_count": bulk_queryset.count(),
        "bulk_municipality": bulk_municipality,
        "bulk_sector": bulk_sector,
        "default_bulk_subject": default_bulk_subject,
        "default_bulk_message": default_bulk_message,
    }
    return render(request, "reports/admin_dashboard.html", context)


@login_required
@_admin_only()
def toggle_user_active(request: HttpRequest, user_id: int) -> HttpResponse:
    """Toggle user active/inactive status from users tab."""
    if request.method != "POST":
        return redirect(f"{reverse('dashboard')}?tab=users")

    user = get_object_or_404(User, id=user_id)
    if user.id == request.user.id:
        messages.error(request, "Не можете да го деактивирате сопствениот профил.")
        return redirect(f"{reverse('dashboard')}?tab=users")

    user.is_active = not user.is_active
    user.save(update_fields=["is_active"])

    _write_audit_log(
        request,
        action="toggle_user_active",
        target_model="User",
        target_id=user.id,
        details={"username": user.username, "is_active": user.is_active},
    )
    state_label = "активиран" if user.is_active else "деактивиран"
    messages.success(request, f"Корисникот {user.username} е {state_label}.")
    return redirect(f"{reverse('dashboard')}?tab=users")


@login_required
@_admin_only()
def create_user(request: HttpRequest) -> HttpResponse:
    """Create a new user from users tab."""
    if request.method == "POST":
        form = AdminUserCreateForm(request.POST)
        if form.is_valid():
            user = User.objects.create_user(
                username=form.cleaned_data["username"],
                email=form.cleaned_data["email"],
                password=form.cleaned_data["password"],
            )
            role = form.cleaned_data["role"]
            profile, _ = UserProfile.objects.get_or_create(user=user)
            profile.role = role
            profile.sector = form.cleaned_data.get("sector") if role == "officer" else ""
            profile.municipality = form.cleaned_data.get("municipality") if role == "officer" else ""
            profile.must_change_password = True
            profile.save(update_fields=["role", "sector", "municipality", "must_change_password"])

            if role == "admin":
                user.is_staff = True
                user.save(update_fields=["is_staff"])

            _write_audit_log(
                request,
                action="create_user",
                target_model="User",
                target_id=user.id,
                details={
                    "username": user.username,
                    "role": role,
                    "sector": profile.sector,
                    "municipality": profile.municipality,
                    "must_change_password": True,
                },
            )
            messages.success(request, f"Корисникот {user.username} е успешно додаден.")
        else:
            error_list = []
            for field_name, field_errors in form.errors.items():
                for field_error in field_errors:
                    if field_name == "__all__":
                        error_list.append(str(field_error))
                    else:
                        error_list.append(f"{field_name}: {field_error}")

            error_message = error_list[0] if error_list else "Неуспешно додавање корисник. Проверете ги полињата."
            messages.error(request, error_message)
    return redirect(f"{reverse('dashboard')}?tab=users")


@login_required
@_admin_only()
def update_user(request: HttpRequest, user_id: int) -> HttpResponse:
    """Update a user's role and worker assignment from the users tab."""
    if request.method != "POST":
        return redirect(f"{reverse('dashboard')}?tab=users")

    user = get_object_or_404(User, id=user_id)
    profile, _ = UserProfile.objects.get_or_create(user=user)
    form = AdminUserUpdateForm(request.POST)

    if form.is_valid():
        role = form.cleaned_data["role"]
        profile.role = role
        profile.sector = form.cleaned_data.get("sector") if role == "officer" else ""
        profile.municipality = form.cleaned_data.get("municipality") if role == "officer" else ""
        profile.save(update_fields=["role", "sector", "municipality"])

        user.is_staff = role == "admin" or user.is_superuser
        user.save(update_fields=["is_staff"])

        _write_audit_log(
            request,
            action="update_user_assignment",
            target_model="User",
            target_id=user.id,
            details={
                "username": user.username,
                "role": role,
                "sector": profile.sector,
                "municipality": profile.municipality,
            },
        )
        messages.success(request, f"Корисникот {user.username} е ажуриран.")
    else:
        first_error = next(iter(form.errors.values()))[0] if form.errors else "Неуспешно ажурирање."
        messages.error(request, first_error)

    return redirect(f"{reverse('dashboard')}?tab=users")


@login_required
@_admin_only()
def delete_user(request: HttpRequest, user_id: int) -> HttpResponse:
    """Delete a user from users tab."""
    if request.method == "POST":
        user = get_object_or_404(User, id=user_id)
        if user.id == request.user.id:
            messages.error(request, "Не можете да се избришете сами себе.")
            return redirect(f"{reverse('dashboard')}?tab=users")

        username = user.username
        user.delete()
        _write_audit_log(
            request,
            action="delete_user",
            target_model="User",
            target_id=user_id,
            details={"username": username},
        )
        messages.success(request, f"Корисникот {username} е избришан.")
    return redirect(f"{reverse('dashboard')}?tab=users")


@login_required
@_admin_only()
def create_category(request: HttpRequest) -> HttpResponse:
    """Create a report category from settings tab."""
    if request.method == "POST":
        payload = request.POST.copy()
        name = (payload.get("name") or "").strip()
        if name and not payload.get("key"):
            payload["key"] = _build_unique_key(ReportCategory, name)

        form = ReportCategoryForm(payload)
        if form.is_valid():
            category = form.save()
            _write_audit_log(
                request,
                action="create_category",
                target_model="ReportCategory",
                target_id=category.id,
                details={"key": category.key, "name": category.name},
            )
            messages.success(request, "Категоријата е успешно додадена.")
        else:
            messages.error(request, "Неуспешно додавање категорија. Проверете ги полињата.")
    return redirect(f"{reverse('dashboard')}?tab=settings")


@login_required
@_admin_only()
def update_category(request: HttpRequest, category_id: int) -> HttpResponse:
    """Update a report category from settings tab."""
    if request.method == "POST":
        category = get_object_or_404(ReportCategory, id=category_id)
        form = ReportCategoryForm(request.POST, instance=category)
        if form.is_valid():
            category = form.save()
            _write_audit_log(
                request,
                action="update_category",
                target_model="ReportCategory",
                target_id=category.id,
                details={"key": category.key, "name": category.name, "is_active": category.is_active},
            )
            messages.success(request, "Категоријата е успешно ажурирана.")
        else:
            messages.error(request, "Неуспешно ажурирање на категоријата.")
    return redirect(f"{reverse('dashboard')}?tab=settings")


@login_required
@_admin_only()
def delete_category(request: HttpRequest, category_id: int) -> HttpResponse:
    """Delete a report category from settings tab."""
    if request.method == "POST":
        category = get_object_or_404(ReportCategory, id=category_id)
        name = category.name
        category.delete()
        _write_audit_log(
            request,
            action="delete_category",
            target_model="ReportCategory",
            target_id=category_id,
            details={"name": name},
        )
        messages.success(request, "Категоријата е избришана.")
    return redirect(f"{reverse('dashboard')}?tab=settings")


@login_required
@_admin_only()
def create_sector(request: HttpRequest) -> HttpResponse:
    """Create a sector from settings tab."""
    if request.method == "POST":
        payload = request.POST.copy()
        name = (payload.get("name") or "").strip()
        if name and not payload.get("key"):
            payload["key"] = _build_unique_key(Sector, name)
        if "is_active" not in payload:
            payload["is_active"] = "true"

        form = SectorForm(payload)
        if form.is_valid():
            sector = form.save()
            _write_audit_log(
                request,
                action="create_sector",
                target_model="Sector",
                target_id=sector.id,
                details={"key": sector.key, "name": sector.name},
            )
            messages.success(request, "Секторот е успешно додаден.")
        else:
            messages.error(request, "Неуспешно додавање сектор. Проверете ги полињата.")
    return redirect(f"{reverse('dashboard')}?tab=settings")


@login_required
@_admin_only()
def update_sector(request: HttpRequest, sector_id: int) -> HttpResponse:
    """Update a sector from settings tab."""
    if request.method == "POST":
        sector = get_object_or_404(Sector, id=sector_id)
        form = SectorForm(request.POST, instance=sector)
        if form.is_valid():
            sector = form.save()
            _write_audit_log(
                request,
                action="update_sector",
                target_model="Sector",
                target_id=sector.id,
                details={"key": sector.key, "name": sector.name, "is_active": sector.is_active},
            )
            state_label = "видлив" if sector.is_active else "скриен"
            messages.success(request, f"Секторот „{sector.name}“ сега е {state_label}.")
        else:
            messages.error(request, "Неуспешно ажурирање на секторот.")
    return redirect(f"{reverse('dashboard')}?tab=settings")


@login_required
@_admin_only()
def delete_sector(request: HttpRequest, sector_id: int) -> HttpResponse:
    """Delete a sector from settings tab."""
    if request.method == "POST":
        sector = get_object_or_404(Sector, id=sector_id)
        name = sector.name
        sector.delete()
        _write_audit_log(
            request,
            action="delete_sector",
            target_model="Sector",
            target_id=sector_id,
            details={"name": name},
        )
        messages.success(request, "Секторот е избришан.")
    return redirect(f"{reverse('dashboard')}?tab=settings")


@login_required
@_admin_only()
def export_reports_csv(request: HttpRequest) -> HttpResponse:
    """Export filtered reports to CSV from dashboard."""
    queryset = _filtered_admin_reports(request)
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="reports_export.csv"'

    writer = csv.writer(response)
    writer.writerow(REPORT_EXPORT_COLUMNS)
    for report in queryset:
        writer.writerow(_format_report_row(report))

    _write_audit_log(
        request,
        action="export_reports_csv",
        target_model="Report",
        target_id=None,
        details={"count": queryset.count(), "filters": request.GET.dict()},
    )
    return response


@login_required
@_admin_only()
def export_reports_excel(request: HttpRequest) -> HttpResponse:
    """Export filtered reports to XLSX from dashboard."""
    if not OPENPYXL_AVAILABLE:
        messages.error(request, "Excel извозот бара openpyxl. Користете CSV извоз.")
        return redirect(f"{reverse('dashboard')}?tab=analytics")

    queryset = _filtered_admin_reports(request)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Reports"
    sheet.append(REPORT_EXPORT_COLUMNS)
    for report in queryset:
        sheet.append(_format_report_row(report))

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="reports_export.xlsx"'
    workbook.save(response)

    _write_audit_log(
        request,
        action="export_reports_excel",
        target_model="Report",
        target_id=None,
        details={"count": queryset.count(), "filters": request.GET.dict()},
    )
    return response


@login_required
@_admin_only()
def import_reports(request: HttpRequest) -> HttpResponse:
    """Validate and import CSV/XLSX rows from dashboard."""
    if request.method != "POST":
        return redirect(f"{reverse('dashboard')}?tab=analytics")

    uploaded_file = request.FILES.get("file")
    if not uploaded_file:
        messages.error(request, "Изберете CSV или XLSX датотека за импорт.")
        return redirect(f"{reverse('dashboard')}?tab=analytics")

    try:
        rows = _read_import_rows(uploaded_file)
        inserted, skipped_duplicates, invalid_rows = _import_reports_from_rows(rows, request.user)
    except ValueError as exc:
        messages.error(request, str(exc))
        return redirect(f"{reverse('dashboard')}?tab=analytics")

    if inserted:
        messages.success(request, f"Импортирани се {inserted} валидни пријави.")
    if skipped_duplicates:
        preview = ", ".join(skipped_duplicates[:8])
        extra = "..." if len(skipped_duplicates) > 8 else ""
        messages.warning(request, f"Прескокнати дупликат ID: {preview}{extra}")
    if invalid_rows:
        preview = "; ".join(f"ред {item['row']}: {item['reason']}" for item in invalid_rows[:8])
        extra = " ..." if len(invalid_rows) > 8 else ""
        messages.error(request, f"Невалидни редови: {preview}{extra}")
    if not inserted and not skipped_duplicates and not invalid_rows:
        messages.info(request, "Датотеката не содржи редови за импорт.")

    _write_audit_log(
        request,
        action="import_reports",
        target_model="Report",
        target_id=None,
        details={
            "inserted": inserted,
            "duplicates": len(skipped_duplicates),
            "invalid": len(invalid_rows),
            "filename": uploaded_file.name,
        },
    )
    return redirect(f"{reverse('dashboard')}?tab=analytics")


@login_required
def submit_report(request):
    """Render and process report submission form."""
    if request.method == "POST":
        form = ReportSubmissionForm(request.POST, request.FILES)
        if form.is_valid():
            if _is_submission_rate_limited(request.user):
                messages.error(
                    request,
                    "Го достигнавте дневниот лимит од 10 пријави за 24 часа. Обидете се повторно утре.",
                )
                return render(
                    request,
                    "reports/submit_report.html",
                    {
                        "form": form,
                        "remaining_quota": _remaining_reports_quota(request.user),
                        "max_reports_per_24h": MAX_REPORTS_PER_24H,
                    },
                    status=429,
                )

            report = form.save(commit=False)
            report.citizen = request.user

            duplicate = find_potential_duplicate(
                description=report.description,
                latitude=float(report.latitude),
                longitude=float(report.longitude),
            )
            if duplicate is not None:
                report.is_duplicate = True
                report.duplicate_of = duplicate
                report.duplicate_verdict = "pending"
                messages.warning(
                    request,
                    f"Можно е оваа пријава да е дупликат на пријава #{duplicate.pk}. Администратор ќе одлучи дали навистина е дупликат.",
                )

            if getattr(settings, "AI_CLASSIFICATION_ENABLED", False):
                try:
                    _apply_ai_classification(report)
                except Exception:
                    report.status = "unclassified"
                    report.category = "other"
                    report.priority = "normal"
                    report.sector = "admin"
                    report.ai_processed = False
                    report.status_changed_at = timezone.now()
            else:
                if not report.category:
                    report.category = "other"
                if not report.priority:
                    report.priority = "normal"
                category_to_sector = {
                    "infrastructure": "infrastructure",
                    "utilities": "utilities",
                    "safety": "safety",
                    "health": "health",
                    "other": "admin",
                }
                report.sector = category_to_sector.get(report.category, "admin")

            report.save()
            _log_status_transition(report, None, report.status, changed_by=request.user, note="Креирана пријава")
            send_report_created_email(report)

            messages.success(request, "Вашата пријава е успешно поднесена.")
            return redirect("report_detail", report_id=report.id)
        messages.error(request, "Ве молиме поправете ги грешките во формата.")
    else:
        form = ReportSubmissionForm()

    return render(
        request,
        "reports/submit_report.html",
        {
            "form": form,
            "remaining_quota": _remaining_reports_quota(request.user),
            "max_reports_per_24h": MAX_REPORTS_PER_24H,
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
def reports_api(request):
    """GET filtered reports, POST create new report with AI classification."""
    if request.method == "POST":
        payload = _parse_json_body(request)
        if payload is None:
            form = ReportCreateForm(request.POST, request.FILES)
        elif not payload:
            return JsonResponse({"errors": {"body": ["Invalid JSON body."]}}, status=400)
        else:
            form = ReportCreateForm(payload)

        if not form.is_valid():
            return JsonResponse({"errors": form.errors}, status=400)

        if _is_submission_rate_limited(request.user):
            return JsonResponse(
                {
                    "detail": "Daily limit reached: max 10 reports per 24 hours. Please try again tomorrow.",
                    "limit": MAX_REPORTS_PER_24H,
                    "window_hours": REPORT_WINDOW_HOURS,
                },
                status=429,
            )

        report = Report(
            citizen=request.user,
            description=form.cleaned_data["description"],
            latitude=form.cleaned_data["latitude"],
            longitude=form.cleaned_data["longitude"],
            image=form.cleaned_data.get("image"),
        )
        _apply_ai_classification(report)
        report.save()
        _log_status_transition(report, None, report.status, changed_by=request.user, note="Креирана пријава")
        send_report_created_email(report)
        return JsonResponse(_serialize_report(report), status=201)

    queryset = _visible_reports_for_user(request)
    queryset = _filter_queryset(request, queryset)

    page = max(int(request.GET.get("page", 1) or 1), 1)
    per_page = min(max(int(request.GET.get("per_page", 20) or 20), 1), 100)
    paginator = Paginator(queryset, per_page)
    page_obj = paginator.get_page(page)

    return JsonResponse(
        {
            "count": paginator.count,
            "num_pages": paginator.num_pages,
            "current_page": page_obj.number,
            "results": [_serialize_report(report) for report in page_obj.object_list],
        }
    )


@login_required
@require_GET
def my_reports(request):
    if request.user.is_authenticated:
        base_qs = Report.objects.filter(citizen=request.user)
    else:
        base_qs = Report.objects.none()

    total_count = base_qs.count()
    new_count = base_qs.filter(status="new").count()
    in_progress_count = base_qs.filter(status="in_progress").count()
    done_count = base_qs.filter(status="resolved").count()
    duplicate_count = base_qs.filter(duplicate_verdict__in=["pending", "confirmed"]).count()

    qs = base_qs.select_related("duplicate_of", "duplicate_of__citizen")

    category = request.GET.get("category", "")
    priority = request.GET.get("priority", "")
    status = request.GET.get("status", "")

    if category:
        qs = qs.filter(category=category)
    if priority:
        qs = qs.filter(priority=priority)
    if status:
        qs = qs.filter(status=status)

    map_points = [
        {
            "id": r.id,
            "lat": float(r.latitude),
            "lng": float(r.longitude),
            "category": r.get_category_display(),
            "status": r.get_status_display(),
            "status_key": r.status,
        }
        for r in qs
    ]

    return render(
        request,
        "reports/my_reports.html",
        {
            "reports": qs,
            "map_points": map_points,
            "total_count": total_count,
            "new_count": new_count,
            "in_progress_count": in_progress_count,
            "done_count": done_count,
            "duplicate_count": duplicate_count,
            "category_choices": Report.CATEGORY_CHOICES,
            "priority_choices": Report.PRIORITY_CHOICES,
            "status_choices": Report.STATUS_CHOICES,
            "selected_category": category,
            "selected_priority": priority,
            "selected_status": status,
        },
    )


@login_required
@require_GET
def report_detail(request, report_id: int):
    """Show full details for one report (owner/officer same sector/admin)."""
    report = get_object_or_404(
        Report.objects.select_related("citizen", "duplicate_of", "duplicate_of__citizen"),
        pk=report_id,
    )
    if not _can_view_report(request.user, report):
        return JsonResponse({"detail": "Not allowed."}, status=403)
    can_view_duplicate_original = bool(
        report.duplicate_of_id and _can_view_report(request.user, report.duplicate_of)
    )
    status_labels = dict(Report.STATUS_CHOICES)
    timeline = []
    for event in report.status_history.select_related("changed_by").all():
        timeline.append(
            {
                "from_status": event.from_status,
                "to_status": event.to_status,
                "from_status_label": status_labels.get(event.from_status, event.from_status),
                "to_status_label": status_labels.get(event.to_status, event.to_status),
                "changed_by": event.changed_by,
                "changed_at": event.changed_at,
                "note": event.note,
            }
        )

    if not timeline:
        timeline = [
            {
                "from_status": "",
                "to_status": report.status,
                "from_status_label": "",
                "to_status_label": status_labels.get(report.status, report.status),
                "changed_by": None,
                "changed_at": report.created_at,
                "note": "Креирана пријава",
            }
        ]
    return render(
        request,
        "reports/report_detail.html",
        {
            "report": report,
            "timeline": timeline,
            "can_view_duplicate_original": can_view_duplicate_original,
        },
    )


@login_required
@require_http_methods(["POST"])
def withdraw_report(request, report_id: int):
    """Allow citizens to withdraw their own report."""
    report = get_object_or_404(Report, pk=report_id)
    if report.citizen_id != request.user.id:
        return JsonResponse({"detail": "Only owner can withdraw this report."}, status=403)

    if report.status not in {"new", "unclassified"}:
        messages.warning(request, "Повлекување е дозволено само за пријави со статус Нова или Некласифицирана.")
        return redirect("my_reports")

    old_status = report.status
    report.status = "withdrawn"
    report.status_changed_at = timezone.now()
    report.save(update_fields=["status", "status_changed_at", "updated_at"])
    _log_status_transition(report, old_status, report.status, changed_by=request.user, note="Повлечена од корисник")
    send_report_status_changed_email(report, old_status, report.status)
    messages.success(request, f"Пријавата #{report.id} е повлечена.")
    return redirect("my_reports")


@login_required
def update_report_status(request, report_id: int):
    """Officer status update endpoint; accepts PATCH or POST."""
    if request.method not in {"PATCH", "POST"}:
        return HttpResponseNotAllowed(["PATCH", "POST"])

    report = get_object_or_404(Report, pk=report_id)

    if not _is_officer(request.user):
        return JsonResponse({"detail": "Only officers can update status."}, status=403)

    profile = UserProfile.objects.filter(user=request.user).first()
    if not profile or not profile.sector or profile.sector != report.sector:
        return JsonResponse({"detail": "You can only edit reports from your own sector."}, status=403)
    if profile.municipality and profile.municipality != report.municipality:
        return JsonResponse({"detail": "You can only edit reports from your assigned municipality."}, status=403)

    payload = _parse_json_body(request) if request.method == "PATCH" else request.POST
    if payload is None:
        payload = {}

    requested_status = (payload.get("status") or "").strip()
    valid_statuses = {value for value, _ in Report.STATUS_CHOICES}
    if requested_status and requested_status not in valid_statuses:
        return JsonResponse({"errors": {"status": ["Invalid status."]}}, status=400)

    old_status = report.status
    update_fields = []

    if requested_status:
        report.status = requested_status
        report.status_changed_at = timezone.now()
        update_fields.extend(["status", "status_changed_at"])

    if "internal_note" in payload:
        report.internal_note = payload.get("internal_note") or ""
        update_fields.append("internal_note")

    report.assigned_officer = request.user
    update_fields.append("assigned_officer")

    if update_fields:
        report.save(update_fields=update_fields)

    if requested_status and old_status != report.status:
        _log_status_transition(
            report,
            old_status,
            report.status,
            changed_by=request.user,
            note="Промена од службеник",
        )
        html_email_sent = send_report_status_changed_email(report, old_status, report.status)
        notification = send_status_change_email(report)
        email_sent = bool(html_email_sent or (notification and notification.status == "sent"))

    payload = _serialize_report(report)
    payload["internal_note"] = report.internal_note
    if requested_status and old_status != report.status:
        payload["status_changed"] = True
        payload["old_status"] = old_status
        payload["old_status_label"] = dict(Report.STATUS_CHOICES).get(old_status, old_status)
        payload["email_sent"] = email_sent
        payload["email_recipient"] = getattr(report.citizen, "email", "") or ""
    else:
        payload["status_changed"] = False
        payload["email_sent"] = False
    return JsonResponse(payload)


@login_required
@require_http_methods(["POST"])
def reassign_report_sector(request, report_id: int):
    """Allow an officer to send a wrongly assigned report to another sector."""
    if not _is_officer(request.user):
        return JsonResponse({"detail": "Only officers can reassign reports."}, status=403)

    report = get_object_or_404(Report, pk=report_id)
    profile = UserProfile.objects.filter(user=request.user).first()
    if not profile or not profile.sector or profile.sector != report.sector:
        return JsonResponse({"detail": "You can only reassign reports from your own sector."}, status=403)
    if profile.municipality and profile.municipality != report.municipality:
        return JsonResponse({"detail": "You can only reassign reports from your assigned municipality."}, status=403)

    new_sector = (request.POST.get("sector") or "").strip()
    valid_sectors = set(Sector.objects.filter(is_active=True).values_list("key", flat=True)) | {
        value for value, _ in Report.SECTOR_CHOICES
    }
    if not new_sector or new_sector not in valid_sectors:
        return JsonResponse({"errors": {"sector": ["Invalid sector."]}}, status=400)
    if new_sector == report.sector:
        return JsonResponse({"errors": {"sector": ["Choose a different sector."]}}, status=400)

    old_sector = report.sector
    report.sector = new_sector
    report.assigned_officer = None
    report.save(update_fields=["sector", "assigned_officer", "updated_at"])

    _write_audit_log(
        request,
        action="reassign_report_sector",
        target_model="Report",
        target_id=report.id,
        details={"old_sector": old_sector, "new_sector": new_sector},
    )
    notify_report_reassigned(report, old_sector, reassigned_by=request.user)

    return JsonResponse(
        {
            "id": report.id,
            "old_sector": old_sector,
            "new_sector": report.sector,
            "detail": "Report reassigned.",
        }
    )


@login_required
@_admin_only()
@require_http_methods(["POST"])
def admin_classify_report(request, report_id: int):
    """Admin endpoint to classify unclassified reports."""
    report = get_object_or_404(Report, pk=report_id)
    
    category = (request.POST.get("category") or "").strip()
    priority = (request.POST.get("priority") or "").strip()
    sector = (request.POST.get("sector") or "").strip()
    
    valid_categories = {value for value, _ in Report.CATEGORY_CHOICES}
    valid_priorities = {value for value, _ in Report.PRIORITY_CHOICES}
    valid_sectors = {value for value, _ in Report.SECTOR_CHOICES}
    
    errors = {}
    if category and category not in valid_categories:
        errors["category"] = "Invalid category"
    if priority and priority not in valid_priorities:
        errors["priority"] = "Invalid priority"
    if sector and sector not in valid_sectors:
        errors["sector"] = "Invalid sector"
    
    if errors:
        return JsonResponse({"errors": errors}, status=400)
    
    update_fields = []
    old_category = report.category
    old_priority = report.priority
    old_sector = report.sector
    
    if category and category != "other":
        report.category = category
        update_fields.append("category")
    
    if priority:
        report.priority = priority
        update_fields.append("priority")
    
    if sector:
        report.sector = sector
        update_fields.append("sector")
    
    # If classified (no longer "other"), update status if needed
    if category and category != "other" and report.status == "unclassified":
        report.status = "new"
        update_fields.append("status")
    
    if update_fields:
        report.save(update_fields=update_fields)
        _write_audit_log(
            request,
            action="classify_report",
            target_model="Report",
            target_id=report.id,
            details={
                "old_category": old_category,
                "new_category": report.category,
                "old_priority": old_priority,
                "new_priority": report.priority,
                "old_sector": old_sector,
                "new_sector": report.sector,
            },
        )
        notify_report_classified(report, classified_by=request.user)
        messages.success(request, "Извештајот е успешно класифициран.")
    
    return redirect(f"{reverse('dashboard')}?tab=unclassified")


@login_required
@_admin_only()
@require_http_methods(["POST"])
def review_duplicate_report(request, report_id: int):
    """Admin confirms or rejects automatic duplicate suggestion."""
    report = get_object_or_404(Report, pk=report_id)
    if report.duplicate_verdict != "pending":
        messages.error(request, "Оваа пријава не е во редица за преглед на дупликат.")
        return redirect(f"{reverse('dashboard')}?tab=duplicates")

    action = (request.POST.get("action") or "").strip()
    if action == "confirm":
        report.duplicate_verdict = "confirmed"
        report.is_duplicate = True
        report.save(update_fields=["duplicate_verdict", "is_duplicate", "updated_at"])
        _write_audit_log(
            request,
            action="duplicate_verdict_confirm",
            target_model="Report",
            target_id=report.id,
            details={
                "duplicate_of_id": report.duplicate_of_id,
            },
        )
        messages.success(
            request,
            f"ПРЈ-{report.id} е означена како дупликат на ПРЈ-{report.duplicate_of_id}.",
        )
    elif action == "reject":
        old_dup_id = report.duplicate_of_id
        report.duplicate_verdict = "rejected"
        report.is_duplicate = False
        report.duplicate_of = None
        report.save(update_fields=["duplicate_verdict", "is_duplicate", "duplicate_of", "updated_at"])
        _write_audit_log(
            request,
            action="duplicate_verdict_reject",
            target_model="Report",
            target_id=report.id,
            details={"previous_duplicate_of_id": old_dup_id},
        )
        messages.success(
            request,
            f"ПРЈ-{report.id} е задржана како посебна пријава (не е дупликат).",
        )
    else:
        messages.error(request, "Непозната акција.")

    return redirect(f"{reverse('dashboard')}?tab=duplicates")


@login_required
def map_view(request):
    """Render interactive map page with report filters."""
    active_sector_choices = list(Sector.objects.filter(is_active=True).values_list('key', 'name'))
    context = {
        "category_choices": Report.CATEGORY_CHOICES,
        "status_choices": Report.STATUS_CHOICES,
        "sector_choices": active_sector_choices,
        "priority_choices": Report.PRIORITY_CHOICES,
    }
    return render(request, "reports/map.html", context)


@login_required
def reports_json(request):
    """Return reports as JSON for AJAX-based Leaflet map rendering."""
    queryset = _filter_queryset(request, _visible_reports_for_user(request).order_by("-created_at"))

    status_labels = dict(Report.STATUS_CHOICES)
    category_labels = dict(Report.CATEGORY_CHOICES)
    priority_labels = dict(Report.PRIORITY_CHOICES)
    sector_labels = dict(Report.SECTOR_CHOICES)

    data = [
        {
            "id": report.pk,
            "description": report.description,
            "status": report.status,
            "status_label": status_labels.get(report.status, report.status),
            "category": report.category,
            "category_label": category_labels.get(report.category, report.category),
            "priority": report.priority,
            "priority_label": priority_labels.get(report.priority, report.priority),
            "sector": report.sector,
            "sector_label": sector_labels.get(report.sector, report.sector),
            "municipality": report.municipality or "",
            "lat": float(report.latitude),
            "lng": float(report.longitude),
            "created_at": report.created_at.isoformat(),
            "detail_url": reverse("report_detail", args=[report.id]),
        }
        for report in queryset
    ]

    return JsonResponse({"results": data})


@login_required
@require_GET
def heatmap_data(request):
    """Return aggregated heatmap points as {lat,lng,weight}."""
    queryset = _visible_reports_for_user(request)
    buckets = {}

    for report in queryset:
        lat = round(_as_float(report.latitude) or 0.0, 3)
        lng = round(_as_float(report.longitude) or 0.0, 3)
        key = (lat, lng)
        buckets[key] = buckets.get(key, 0) + 1

    points = [{"lat": lat, "lng": lng, "weight": weight} for (lat, lng), weight in buckets.items()]
    return JsonResponse(points, safe=False)


@login_required
def officer_panel(request):
    if not _is_officer(request.user):
        return redirect("dashboard")

    profile = UserProfile.objects.filter(user=request.user).first()
    sector = profile.sector if profile else None
    reports = Report.objects.filter(sector=sector).select_related("duplicate_of").order_by("-created_at")
    if profile and profile.municipality:
        reports = reports.filter(municipality=profile.municipality)

    status_filter = request.GET.get("status")
    if status_filter:
        reports = reports.filter(status=status_filter)

    priority_filter = request.GET.get("priority")
    if priority_filter:
        reports = reports.filter(priority=priority_filter)

    active_sector_choices = list(Sector.objects.filter(is_active=True).values_list("key", "name"))
    destination_sector_choices = [
        (key, name) for key, name in active_sector_choices if key != sector
    ]

    return render(
        request,
        "reports/officer_panel.html",
        {
            "reports": reports,
            "sector": sector,
            "municipality": profile.municipality if profile else "",
            "status_choices": Report.STATUS_CHOICES,
            "priority_choices": Report.PRIORITY_CHOICES,
            "category_choices": Report.CATEGORY_CHOICES,
            "sector_choices": active_sector_choices,
            "destination_sector_choices": destination_sector_choices,
            "municipality_choices": MUNICIPALITY_CHOICES,
        },
    )


@login_required
def search_page(request):
    """Search visible reports, with keyword matching report descriptions."""
    filters = _build_report_filters(request)
    opshtina = request.GET.get("opshtina", "").strip()
    if opshtina:
        filters &= Q(municipality=opshtina)

    queryset = _visible_reports_for_user(request).filter(filters).order_by("-created_at")
    sort = request.GET.get("sort", "date")
    sort_map = {
        "date": "-created_at",
        "priority": "priority",
        "status": "status",
    }
    queryset = queryset.order_by(sort_map.get(sort, "-created_at"))
    paginator = Paginator(queryset, 20)
    page_number = request.GET.get("page", 1)
    try:
        page_obj = paginator.page(page_number)
    except (PageNotAnInteger, EmptyPage):
        page_obj = paginator.page(1)

    active_sector_choices = list(Sector.objects.filter(is_active=True).values_list('key', 'name'))
    active_category_choices = list(ReportCategory.objects.filter(is_active=True).values_list('key', 'name'))
    
    return render(
        request,
        "reports/search_results.html",
        {
            "reports_page": page_obj,
            "query": request.GET,
            "status_choices": Report.STATUS_CHOICES,
            "priority_choices": Report.PRIORITY_CHOICES,
            "sector_choices": active_sector_choices,
            "category_choices": active_category_choices,
            "municipalities": MUNICIPALITY_CHOICES,
        },
    )
